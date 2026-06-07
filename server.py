import asyncio
import json
import os
import io
import re
import httpx
from contextlib import asynccontextmanager
from fastapi import FastAPI, UploadFile, Form, HTTPException, Header
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import psycopg2
import psycopg2.extras
import xlrd
import openpyxl

DATABASE_URL = os.environ.get("DATABASE_URL")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "vm2026")
MATTERMOST_WEBHOOK = os.environ.get("MATTERMOST_WEBHOOK", "https://jesper-matter.synology.me/hooks/e81ai3c98fyjdepjxkq7nx3qkr")
POLL_INTERVAL = int(os.environ.get("POLL_INTERVAL", "60"))  # seconds

JESPER_TIPS = {1:{"home":2,"away":0},2:{"home":1,"away":0},3:{"home":2,"away":0},4:{"home":2,"away":1},5:{"home":0,"away":2},6:{"home":2,"away":0},7:{"home":0,"away":2},8:{"home":0,"away":1},9:{"home":4,"away":0},10:{"home":2,"away":0},11:{"home":0,"away":1},12:{"home":2,"away":1},13:{"home":4,"away":0},14:{"home":2,"away":1},15:{"home":0,"away":2},16:{"home":2,"away":1},17:{"home":2,"away":1},18:{"home":0,"away":2},19:{"home":2,"away":0},20:{"home":2,"away":0},21:{"home":3,"away":0},22:{"home":2,"away":0},23:{"home":1,"away":0},24:{"home":0,"away":2},25:{"home":1,"away":0},26:{"home":2,"away":0},27:{"home":3,"away":0},28:{"home":2,"away":0},29:{"home":1,"away":0},30:{"home":0,"away":1},31:{"home":3,"away":0},32:{"home":1,"away":1},33:{"home":2,"away":0},34:{"home":2,"away":0},35:{"home":2,"away":0},36:{"home":0,"away":2},37:{"home":4,"away":0},38:{"home":2,"away":0},39:{"home":4,"away":0},40:{"home":0,"away":2},41:{"home":2,"away":0},42:{"home":3,"away":0},43:{"home":1,"away":1},44:{"home":1,"away":1},45:{"home":4,"away":0},46:{"home":3,"away":0},47:{"home":0,"away":3},48:{"home":2,"away":0},49:{"home":1,"away":1},50:{"home":2,"away":1},51:{"home":0,"away":3},52:{"home":3,"away":0},53:{"home":0,"away":2},54:{"home":1,"away":1},55:{"home":0,"away":4},56:{"home":0,"away":3},57:{"home":1,"away":1},58:{"home":0,"away":3},59:{"home":1,"away":2},60:{"home":1,"away":1},61:{"home":1,"away":2},62:{"home":3,"away":0},63:{"home":0,"away":1},64:{"home":0,"away":2},65:{"home":1,"away":1},66:{"home":0,"away":2},67:{"home":0,"away":3},68:{"home":2,"away":0},69:{"home":0,"away":2},70:{"home":1,"away":1},71:{"home":0,"away":1},72:{"home":0,"away":3}}

# match_num → "Home - Away" (used to match ESPN team names)
MATCHES = {
    1:"Mexico - South Africa", 2:"South Korea - Czechia", 3:"Canada - Bosnia-Herzegovina",
    4:"USA - Paraguay", 5:"Qatar - Switzerland", 6:"Brazil - Morocco",
    7:"Haiti - Scotland", 8:"Australia - Türkiye", 9:"Germany - Curaçao",
    10:"Netherlands - Japan", 11:"Ivory Coast - Ecuador", 12:"Sweden - Tunisia",
    13:"Spain - Cape Verde", 14:"Belgium - Egypt", 15:"Saudi Arabia - Uruguay",
    16:"Iran - New Zealand", 17:"France - Senegal", 18:"Iraq - Norway",
    19:"Argentina - Algeria", 20:"Austria - Jordan", 21:"Portugal - Congo DR",
    22:"England - Croatia", 23:"Ghana - Panama", 24:"Uzbekistan - Colombia",
    25:"Czechia - South Africa", 26:"Switzerland - Bosnia-Herzegovina", 27:"Canada - Qatar",
    28:"Mexico - South Korea", 29:"USA - Australia", 30:"Scotland - Morocco",
    31:"Brazil - Haiti", 32:"Türkiye - Paraguay", 33:"Netherlands - Sweden",
    34:"Germany - Ivory Coast", 35:"Ecuador - Curaçao", 36:"Tunisia - Japan",
    37:"Spain - Saudi Arabia", 38:"Belgium - Iran", 39:"Uruguay - Cape Verde",
    40:"New Zealand - Egypt", 41:"Argentina - Austria", 42:"France - Iraq",
    43:"Norway - Senegal", 44:"Jordan - Algeria", 45:"Portugal - Uzbekistan",
    46:"England - Ghana", 47:"Panama - Croatia", 48:"Colombia - Congo DR",
    49:"Switzerland - Canada", 50:"Bosnia-Herzegovina - Qatar", 51:"Scotland - Brazil",
    52:"Morocco - Haiti", 53:"Czechia - Mexico", 54:"South Africa - South Korea",
    55:"Curaçao - Ivory Coast", 56:"Ecuador - Germany", 57:"Japan - Sweden",
    58:"Tunisia - Netherlands", 59:"Türkiye - USA", 60:"Paraguay - Australia",
    61:"Norway - France", 62:"Senegal - Iraq", 63:"Cape Verde - Saudi Arabia",
    64:"Uruguay - Spain", 65:"Egypt - Iran", 66:"New Zealand - Belgium",
    67:"Panama - England", 68:"Croatia - Ghana", 69:"Colombia - Portugal",
    70:"Congo DR - Uzbekistan", 71:"Algeria - Austria", 72:"Jordan - Argentina",
}


# ── Database ───────────────────────────────────────────────────────────────────

def get_db():
    return psycopg2.connect(DATABASE_URL)


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS participants (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            tips JSONB NOT NULL,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS posted_matches (
            match_num INTEGER PRIMARY KEY,
            home_score INTEGER,
            away_score INTEGER,
            posted_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS match_results (
            match_num INTEGER PRIMARY KEY,
            home_score INTEGER,
            away_score INTEGER,
            status TEXT DEFAULT 'upcoming',
            minute TEXT,
            kickoff TEXT,
            home_name TEXT,
            away_name TEXT,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("SELECT COUNT(*) FROM participants")
    if cur.fetchone()[0] == 0:
        tips = {str(k): v for k, v in JESPER_TIPS.items()}
        cur.execute(
            "INSERT INTO participants (name, tips) VALUES (%s, %s)",
            ("Jesper", json.dumps(tips))
        )
    conn.commit()
    cur.close()
    conn.close()


def load_data():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT name, tips FROM participants ORDER BY id")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{"name": r["name"], "tips": r["tips"]} for r in rows]


def get_posted_matches() -> set:
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT match_num FROM posted_matches")
    nums = {row[0] for row in cur.fetchall()}
    cur.close()
    conn.close()
    return nums


def mark_match_posted(match_num: int, home: int, away: int):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO posted_matches (match_num, home_score, away_score) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
        (match_num, home, away)
    )
    conn.commit()
    cur.close()
    conn.close()


# ── ESPN ───────────────────────────────────────────────────────────────────────

def norm(s: str) -> str:
    return re.sub(r'[^a-z]', '', s.lower())


def find_match_num(home_name: str, away_name: str) -> int | None:
    h, a = norm(home_name), norm(away_name)
    for num, teams in MATCHES.items():
        parts = teams.split(' - ')
        if len(parts) < 2:
            continue
        mh, ma = norm(parts[0]), norm(parts[1])
        if (mh in h or h in mh) and (ma in a or a in ma):
            return num
        if (mh in a or a in mh) and (ma in h or h in ma):
            return num
    return None


async def fetch_and_store_results() -> dict:
    """Fetches all WC matches from ESPN, stores in DB, returns {match_num: {...}}."""
    from datetime import datetime, timezone, timedelta

    all_results = {}
    url = "https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard?dates=20260611-20260710&limit=200"
    async with httpx.AsyncClient(timeout=15) as client:
        try:
            r = await client.get(url)
            events = r.json().get("events", [])
        except Exception as e:
            print(f"ESPN fetch error: {e}")
            return all_results

    conn = get_db()
    cur = conn.cursor()

    for ev in events:
        comp = (ev.get("competitions") or [{}])[0]
        espn_status = (comp.get("status") or {}).get("type", {}).get("name", "")
        competitors = comp.get("competitors", [])
        home = next((c for c in competitors if c.get("homeAway") == "home"), None)
        away = next((c for c in competitors if c.get("homeAway") == "away"), None)
        if not home or not away:
            continue

        home_name = (home.get("team") or {}).get("shortDisplayName") or (home.get("team") or {}).get("displayName", "")
        away_name = (away.get("team") or {}).get("shortDisplayName") or (away.get("team") or {}).get("displayName", "")
        match_num = find_match_num(home_name, away_name)
        if not match_num:
            continue

        home_score = int(home.get("score") or 0) if espn_status != "STATUS_SCHEDULED" else None
        away_score = int(away.get("score") or 0) if espn_status != "STATUS_SCHEDULED" else None
        minute = comp.get("status", {}).get("displayClock") if espn_status == "STATUS_IN_PROGRESS" else None

        # Convert kickoff time to Swedish time
        kickoff = None
        try:
            dt = datetime.fromisoformat(ev.get("date", "").replace("Z", "+00:00"))
            swe = dt.astimezone(timezone(timedelta(hours=2)))  # CEST
            kickoff = swe.strftime("%H:%M")
        except Exception:
            pass

        if espn_status == "STATUS_FINAL":
            status = "done"
        elif espn_status == "STATUS_IN_PROGRESS":
            status = "live"
        else:
            status = "upcoming"

        all_results[match_num] = {
            "home": home_score, "away": away_score,
            "status": status, "minute": minute,
            "kickoff": kickoff, "home_name": home_name, "away_name": away_name,
            "label": f"{home_name} {home_score}–{away_score} {away_name}" if status == "done" else None,
        }

        cur.execute("""
            INSERT INTO match_results (match_num, home_score, away_score, status, minute, kickoff, home_name, away_name, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
            ON CONFLICT (match_num) DO UPDATE SET
                home_score = EXCLUDED.home_score, away_score = EXCLUDED.away_score,
                status = EXCLUDED.status, minute = EXCLUDED.minute,
                kickoff = EXCLUDED.kickoff, home_name = EXCLUDED.home_name,
                away_name = EXCLUDED.away_name, updated_at = NOW()
        """, (match_num, home_score, away_score, status, minute, kickoff, home_name, away_name))

    conn.commit()
    cur.close()
    conn.close()
    return all_results


def load_results_from_db() -> dict:
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM match_results")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {r["match_num"]: dict(r) for r in rows}


# ── Scoring & standings ────────────────────────────────────────────────────────

def calc_score(tip_h: int, tip_a: int, real_h: int, real_a: int) -> int:
    if tip_h == real_h and tip_a == real_a:
        return 3
    tip_sign = (tip_h > tip_a) - (tip_h < tip_a)
    real_sign = (real_h > real_a) - (real_h < real_a)
    return 1 if tip_sign == real_sign else 0


def compute_standings(participants: list, results: dict) -> list:
    rows = []
    for p in participants:
        total = exact = sign = played = 0
        for m_str, tip in p["tips"].items():
            r = results.get(int(m_str))
            if not r:
                continue
            played += 1
            s = calc_score(tip["home"], tip["away"], r["home"], r["away"])
            total += s
            if s == 3:
                exact += 1
            elif s == 1:
                sign += 1
        rows.append({"name": p["name"], "total": total, "exact": exact, "sign": sign, "played": played})
    rows.sort(key=lambda x: (-x["total"], -x["exact"], -x["sign"]))
    return rows


# ── Mattermost ─────────────────────────────────────────────────────────────────

def format_message(standings: list, match_label: str, exact: list, sign: list, rank_changes: dict) -> str:
    medals = ["🥇", "🥈", "🥉"]
    lines = [f"### ⚽ {match_label}", ""]

    # Who scored what on this match
    if exact:
        lines.append(f"🎯 **3 poäng:** {', '.join(exact)}")
    if sign:
        lines.append(f"✅ **1 poäng:** {', '.join(sign)}")
    if not exact and not sign:
        lines.append("❌ Ingen fick poäng på denna match")
    lines.append("")

    # Standings
    lines.append("**Ställning**")
    for i, row in enumerate(standings):
        prefix = medals[i] if i < 3 else f"{i+1}."
        name = row["name"]
        change = rank_changes.get(name, 0)
        arrow = f" ↑{change}" if change > 0 else (f" ↓{abs(change)}" if change < 0 else "")
        lines.append(f"{prefix} **{name}**{arrow} — {row['total']}p  _(✓{row['exact']} ~{row['sign']} / {row['played']} matcher)_")

    return "\n".join(lines)


async def post_to_mattermost(text: str) -> bool:
    async with httpx.AsyncClient(timeout=10) as client:
        try:
            r = await client.post(MATTERMOST_WEBHOOK, json={"text": text})
            return r.status_code == 200
        except Exception as e:
            print(f"Mattermost error: {e}")
            return False


# ── Background poller ──────────────────────────────────────────────────────────

async def poll_loop():
    print("Poller started")
    while True:
        try:
            all_results = await fetch_and_store_results()
            finished = {num: info for num, info in all_results.items() if info["status"] == "done"}
            posted = get_posted_matches()
            new_matches = {num: info for num, info in finished.items() if num not in posted}

            if new_matches:
                participants = load_data()
                all_results = {num: {"home": info["home"], "away": info["away"]} for num, info in finished.items()}

                for match_num, info in sorted(new_matches.items()):
                    # Standings BEFORE this match
                    results_before = {n: {"home": r["home"], "away": r["away"]} for n, r in finished.items() if n != match_num}
                    standings_before = compute_standings(participants, results_before)
                    ranks_before = {row["name"]: i for i, row in enumerate(standings_before)}

                    # Standings AFTER this match
                    results_after = {n: {"home": r["home"], "away": r["away"]} for n, r in finished.items() if n <= match_num}
                    standings_after = compute_standings(participants, results_after)
                    ranks_after = {row["name"]: i for i, row in enumerate(standings_after)}

                    # Rank changes (positive = moved up)
                    rank_changes = {name: ranks_before.get(name, 0) - ranks_after[name]
                                    for name in ranks_after}

                    # Who got 3p / 1p on this specific match
                    rh, ra = info["home"], info["away"]
                    exact_scorers, sign_scorers = [], []
                    for p in participants:
                        tip = p["tips"].get(str(match_num))
                        if not tip:
                            continue
                        s = calc_score(tip["home"], tip["away"], rh, ra)
                        if s == 3:
                            exact_scorers.append(p["name"])
                        elif s == 1:
                            sign_scorers.append(p["name"])

                    text = format_message(standings_after, info["label"], exact_scorers, sign_scorers, rank_changes)
                    ok = await post_to_mattermost(text)
                    if ok:
                        mark_match_posted(match_num, info["home"], info["away"])
                        print(f"Posted match {match_num}: {info['label']}")
                    else:
                        print(f"Failed to post match {match_num} to Mattermost")

        except Exception as e:
            print(f"Poll loop error: {e}")

        await asyncio.sleep(POLL_INTERVAL)


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    task = asyncio.create_task(poll_loop())
    yield
    task.cancel()
    try:
        await task
    except asyncio.CancelledError:
        pass


app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Excel parsing ──────────────────────────────────────────────────────────────

def to_num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_tips_from_bytes(file_bytes: bytes, filename: str) -> dict:
    tips = {}

    if filename.lower().endswith(".xls"):
        rb = xlrd.open_workbook(file_contents=file_bytes)
        rs = rb.sheet_by_index(0)
        for r in range(6, 42):
            if r >= rs.nrows:
                break
            n1 = to_num(rs.cell_value(r, 0))
            if n1 is not None and 1 <= n1 <= 36:
                h = to_num(rs.cell_value(r, 6))
                a = to_num(rs.cell_value(r, 7))
                if h is not None and a is not None:
                    tips[str(int(n1))] = {"home": int(h), "away": int(a)}
            n2 = to_num(rs.cell_value(r, 13))
            if n2 is not None and 37 <= n2 <= 72:
                h = to_num(rs.cell_value(r, 19))
                a = to_num(rs.cell_value(r, 20))
                if h is not None and a is not None:
                    tips[str(int(n2))] = {"home": int(h), "away": int(a)}

    elif filename.lower().endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        for r in range(7, 43):
            n1 = to_num(ws.cell(r, 1).value)
            if n1 is not None and 1 <= n1 <= 36:
                h = to_num(ws.cell(r, 7).value)
                a = to_num(ws.cell(r, 8).value)
                if h is not None and a is not None:
                    tips[str(int(n1))] = {"home": int(h), "away": int(a)}
            n2 = to_num(ws.cell(r, 14).value)
            if n2 is not None and 37 <= n2 <= 72:
                h = to_num(ws.cell(r, 20).value)
                a = to_num(ws.cell(r, 21).value)
                if h is not None and a is not None:
                    tips[str(int(n2))] = {"home": int(h), "away": int(a)}

    return tips


# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/")
def serve_html():
    return FileResponse("vm2026.html")


@app.get("/api/results")
def get_results():
    return load_results_from_db()


@app.get("/api/participants")
def get_participants():
    return load_data()


@app.post("/api/participants")
async def add_participant(name: str = Form(...), file: UploadFile = Form(...)):
    filename = file.filename or ""
    if not filename.lower().endswith((".xls", ".xlsx", ".xlsm")):
        raise HTTPException(400, "Endast .xls, .xlsx och .xlsm stöds")

    content = await file.read()
    tips = parse_tips_from_bytes(content, filename)

    if len(tips) == 0:
        raise HTTPException(400, "Hittade inga tips i filen. Kontrollera att mål är ifyllda i kolumn G/H (match 1-36) och T/U (match 37-72).")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM participants WHERE LOWER(name) = LOWER(%s)", (name,))
    existing = cur.fetchone()
    if existing:
        cur.execute(
            "UPDATE participants SET tips = %s WHERE LOWER(name) = LOWER(%s)",
            (json.dumps(tips), name)
        )
        conn.commit()
        cur.close()
        conn.close()
        return {"ok": True, "count": len(tips), "updated": True}

    cur.execute(
        "INSERT INTO participants (name, tips) VALUES (%s, %s)",
        (name, json.dumps(tips))
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"ok": True, "count": len(tips), "updated": False}


@app.delete("/api/participants/{name}")
def delete_participant(name: str, x_admin_password: str = Header(default="")):
    if x_admin_password != ADMIN_PASSWORD:
        raise HTTPException(401, "Fel lösenord")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM participants WHERE LOWER(name) = LOWER(%s)", (name,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        raise HTTPException(404, "Deltagare hittades inte")
    if row[0] == 1:
        cur.close()
        conn.close()
        raise HTTPException(403, "Kan inte ta bort administratören")
    cur.execute("DELETE FROM participants WHERE id = %s", (row[0],))
    conn.commit()
    cur.close()
    conn.close()
    return {"ok": True}
